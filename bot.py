import asyncio
import io
import sqlite3
from datetime import datetime, timedelta

import aiohttp
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart, Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, CallbackQuery, BufferedInputFile,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
)
from aiogram.utils.keyboard import InlineKeyboardBuilder

# ========================= CONFIG =========================

import os
TOKEN = os.getenv("BOT_TOKEN")

# Время ежедневного напоминания (час и минута, по МСК UTC+3)
REMINDER_HOUR = 21
REMINDER_MINUTE = 0


CATEGORIES = [
    ("🍔 Еда", "еда"),
    ("🛒 Магазин", "магазин"),
    ("📱 Подписки", "подписки"),
    ("👕 Одежда", "одежда"),
    ("🏠 Быт", "быт"),
    ("🚌 Транспорт", "транспорт"),
    ("🎉 Развлечения", "развлечения"),
    ("🎮 Игры", "игры"),
    ("💊 Здоровье", "здоровье"),
    ("☕ Кафе", "кафе"),
    ("💻 Техника", "техника"),
    ("🎁 Подарки", "подарки"),
    ("🏋️ Спорт", "спорт"),
    ("💡 Коммуналка", "коммуналка"),
]

CURRENCIES = [
    ("🇺🇸 USD", "USD"),
    ("🇪🇺 EUR", "EUR"),
    ("🇬🇧 GBP", "GBP"),
    ("🇨🇳 CNY", "CNY"),
    ("🇯🇵 JPY", "JPY"),
    ("🇧🇾 BYN", "BYN"),
    ("🇰🇿 KZT", "KZT"),
    ("🇺🇦 UAH", "UAH"),
]

# ========================= DATABASE =========================

def init_db():
    conn = sqlite3.connect("finance.db")
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS expenses(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, amount REAL,
            category TEXT, date TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS incomes(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, amount REAL,
            source TEXT, date TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS goals(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, title TEXT,
            target REAL, current REAL DEFAULT 0
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS budgets(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER, category TEXT,
            amount REAL, month TEXT,
            UNIQUE(user_id, category, month)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS reminder_users(
            user_id INTEGER PRIMARY KEY
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS recurring(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            title TEXT,
            amount REAL,
            category TEXT,
            day_of_month INTEGER,
            last_applied TEXT
        )
    """)


    # Migrations
    for col_sql in [
        "ALTER TABLE incomes ADD COLUMN source TEXT",
    ]:
        try:
            cur.execute(col_sql)
        except Exception:
            pass

    conn.commit()
    conn.close()

def get_conn():
    return sqlite3.connect("finance.db")

def today_date():
    return datetime.now().strftime("%Y-%m-%d")

def current_month():
    return datetime.now().strftime("%Y-%m")

def get_period_stats(uid, days=None):
    conn = get_conn()
    cur = conn.cursor()
    if days:
        start = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        cur.execute("SELECT SUM(amount) FROM incomes WHERE user_id=? AND date>=?", (uid, start))
        income = cur.fetchone()[0] or 0
        cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=? AND date>=?", (uid, start))
        expense = cur.fetchone()[0] or 0
    else:
        cur.execute("SELECT SUM(amount) FROM incomes WHERE user_id=?", (uid,))
        income = cur.fetchone()[0] or 0
        cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=?", (uid,))
        expense = cur.fetchone()[0] or 0
    conn.close()
    return income, expense

def progress_bar(current, target, length=10):
    if target == 0:
        return "░" * length
    filled = min(int((current / target) * length), length)
    return "█" * filled + "░" * (length - filled)

# ========================= FSM STATES =========================

class AddIncome(StatesGroup):
    amount = State()
    source = State()

class AddExpense(StatesGroup):
    amount = State()
    category = State()

class AddGoal(StatesGroup):
    title = State()
    target = State()

class FundGoal(StatesGroup):
    amount = State()

class SetBudget(StatesGroup):
    category = State()
    amount = State()

class CurrencyConvert(StatesGroup):
    from_currency = State()
    amount = State()
    to_currency = State()

class AddRecurring(StatesGroup):
    title = State()
    amount = State()
    category = State()
    day = State()


# ========================= KEYBOARDS =========================

def main_menu_kb():
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Доход"), KeyboardButton(text="➖ Расход")],
            [KeyboardButton(text="🎯 Цели"), KeyboardButton(text="📊 Статистика")],
            [KeyboardButton(text="📋 История"), KeyboardButton(text="⚙️ Управление")],
            [KeyboardButton(text="💱 Валюты"), KeyboardButton(text="📤 Экспорт Excel")],
            [KeyboardButton(text="🔄 Регулярные")],
        ],
        resize_keyboard=True
    )

def stats_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Сегодня", callback_data="stats_today"),
            InlineKeyboardButton(text="📆 Неделя", callback_data="stats_week"),
        ],
        [
            InlineKeyboardButton(text="🗓 Месяц", callback_data="stats_month"),
            InlineKeyboardButton(text="📈 Всё время", callback_data="stats_all"),
        ],
        [
            InlineKeyboardButton(text="🏆 Топ категорий", callback_data="stats_top"),
            InlineKeyboardButton(text="📉 Средний расход", callback_data="stats_avg"),
        ],
        [
            InlineKeyboardButton(text="📊 График расходов", callback_data="stats_chart"),
            InlineKeyboardButton(text="💰 Бюджеты", callback_data="budgets_view"),
        ],
    ])

def history_kb():
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="💰 Доходы", callback_data="hist_incomes"),
            InlineKeyboardButton(text="💸 Расходы", callback_data="hist_expenses"),
        ],
    ])

def categories_kb():
    buttons = []
    row = []
    for label, value in CATEGORIES:
        row.append(InlineKeyboardButton(text=label, callback_data=f"cat_{value}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def management_kb(reminders_on: bool):
    reminder_text = "🔔 Напоминания: ВКЛ ✅" if reminders_on else "🔕 Напоминания: ВЫКЛ"
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="🗑 Удалить доход", callback_data="del_income_list")],
        [InlineKeyboardButton(text="🗑 Удалить расход", callback_data="del_expense_list")],
        [InlineKeyboardButton(text="🗑 Удалить цель", callback_data="del_goal_list")],
        [InlineKeyboardButton(text=reminder_text, callback_data="toggle_reminder")],
        [InlineKeyboardButton(text="⚠️ Очистить все данные", callback_data="clear_all")],
    ])

def goals_action_kb(goal_id):
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💰 Пополнить", callback_data=f"fund_goal_{goal_id}")],
        [InlineKeyboardButton(text="🗑 Удалить", callback_data=f"confirm_del_goal_{goal_id}")],
    ])

def confirm_kb(yes_data, no_data="cancel_del"):
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Да", callback_data=yes_data),
            InlineKeyboardButton(text="❌ Нет", callback_data=no_data),
        ]
    ])

def currencies_kb(prefix="from"):
    buttons = []
    row = []
    for label, code in CURRENCIES:
        row.append(InlineKeyboardButton(text=label, callback_data=f"{prefix}_{code}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    buttons.append([InlineKeyboardButton(text="🇷🇺 RUB", callback_data=f"{prefix}_RUB")])
    return InlineKeyboardMarkup(inline_keyboard=buttons)

def budget_categories_kb():
    buttons = []
    row = []
    for label, value in CATEGORIES:
        row.append(InlineKeyboardButton(text=label, callback_data=f"budget_cat_{value}"))
        if len(row) == 2:
            buttons.append(row)
            row = []
    if row:
        buttons.append(row)
    return InlineKeyboardMarkup(inline_keyboard=buttons)

# ========================= BOT INIT =========================

bot = Bot(token=TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# ========================= /start =========================

@dp.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.clear()
    await message.answer(
        f"👋 Привет, <b>{message.from_user.first_name}</b>!\n\n"
        "Я твой личный <b>финансовый помощник</b> 💼\n\n"
        "Используй кнопки ниже для управления:",
        parse_mode="HTML",
        reply_markup=main_menu_kb()
    )

# ========================= MAIN MENU =========================

@dp.message(F.text == "➕ Доход")
async def menu_income(message: Message, state: FSMContext):
    await state.set_state(AddIncome.amount)
    await message.answer(
        "💰 <b>Добавление дохода</b>\n\nВведите сумму:",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(F.text == "➖ Расход")
async def menu_expense(message: Message, state: FSMContext):
    await state.set_state(AddExpense.amount)
    await message.answer(
        "💸 <b>Добавление расхода</b>\n\nВведите сумму:",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardRemove()
    )

@dp.message(F.text == "📊 Статистика")
async def menu_stats(message: Message):
    await message.answer("📊 <b>Статистика</b>\n\nВыберите период:", parse_mode="HTML", reply_markup=stats_kb())

@dp.message(F.text == "📋 История")
async def menu_history(message: Message):
    await message.answer("📋 <b>История операций</b>\n\nЧто показать?", parse_mode="HTML", reply_markup=history_kb())

@dp.message(F.text == "⚙️ Управление")
async def menu_management(message: Message):
    uid = message.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT user_id FROM reminder_users WHERE user_id=?", (uid,))
    reminders_on = cur.fetchone() is not None
    conn.close()
    await message.answer("⚙️ <b>Управление данными</b>", parse_mode="HTML", reply_markup=management_kb(reminders_on))

@dp.message(F.text == "🎯 Цели")
async def menu_goals(message: Message):
    uid = message.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, title, target, current FROM goals WHERE user_id=?", (uid,))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить цель", callback_data="add_goal")]
        ])
        await message.answer("🎯 У вас пока нет целей.", reply_markup=kb)
        return

    for id_, title, target, current in rows:
        pct = round((current / target) * 100, 1) if target else 0
        bar = progress_bar(current, target)
        done = " 🎉 <b>ВЫПОЛНЕНА!</b>" if current >= target else ""
        text = (
            f"🎯 <b>{title}</b>{done}\n"
            f"{bar} {pct}%\n"
            f"Накоплено: <b>{current:,.0f} ₽</b> из <b>{target:,.0f} ₽</b>\n"
            f"Осталось: <b>{max(target - current, 0):,.0f} ₽</b>"
        )
        await message.answer(text, parse_mode="HTML", reply_markup=goals_action_kb(id_))

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить цель", callback_data="add_goal")]
    ])
    await message.answer("Управление целями:", reply_markup=kb)

# ========================= ADD INCOME FSM =========================

@dp.message(AddIncome.amount)
async def income_get_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму (например: 5000)")
        return
    await state.update_data(amount=amount)
    await state.set_state(AddIncome.source)
    skip_kb = ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="⏭ Пропустить")]],
        resize_keyboard=True, one_time_keyboard=True
    )
    await message.answer(
        f"✅ Сумма: <b>{amount:,.0f} ₽</b>\n\nОткуда доход? (зарплата, фриланс и т.д.)",
        parse_mode="HTML", reply_markup=skip_kb
    )

@dp.message(AddIncome.source)
async def income_get_source(message: Message, state: FSMContext):
    data = await state.get_data()
    amount = data["amount"]
    source = None if message.text.lower() in ("пропустить", "⏭ пропустить") else message.text
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO incomes VALUES(NULL,?,?,?,?)",
                (message.from_user.id, amount, source, today_date()))
    conn.commit()
    conn.close()
    await state.clear()
    src_text = f" ({source})" if source else ""
    await message.answer(
        f"✅ Доход <b>{amount:,.0f} ₽</b>{src_text} добавлен!",
        parse_mode="HTML", reply_markup=main_menu_kb()
    )

# ========================= ADD EXPENSE FSM =========================

@dp.message(AddExpense.amount)
async def expense_get_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму (например: 1500)")
        return
    await state.update_data(amount=amount)
    await state.set_state(AddExpense.category)
    await message.answer(
        f"✅ Сумма: <b>{amount:,.0f} ₽</b>\n\nВыберите категорию:",
        parse_mode="HTML", reply_markup=categories_kb()
    )

@dp.callback_query(F.data.startswith("cat_"), AddExpense.category)
async def expense_get_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data[4:]
    data = await state.get_data()
    amount = data["amount"]
    uid = callback.from_user.id

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO expenses VALUES(NULL,?,?,?,?)",
                (uid, amount, category, today_date()))

    # Check budget
    cur.execute("SELECT amount FROM budgets WHERE user_id=? AND category=? AND month=?",
                (uid, category, current_month()))
    budget_row = cur.fetchone()

    cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=? AND category=? AND date LIKE ?",
                (uid, category, f"{current_month()}%"))
    spent = cur.fetchone()[0] or 0

    conn.commit()
    conn.close()

    await state.clear()
    cat_label = next((l for l, v in CATEGORIES if v == category), category)
    text = f"✅ Расход <b>{amount:,.0f} ₽</b> в категории <b>{cat_label}</b> добавлен!"

    if budget_row:
        budget = budget_row[0]
        pct = round((spent / budget) * 100)
        bar = progress_bar(spent, budget)
        text += f"\n\n💰 Бюджет на месяц: <b>{budget:,.0f} ₽</b>\n{bar} {pct}%\nПотрачено: <b>{spent:,.0f} ₽</b>"
        if spent >= budget:
            text += "\n\n⚠️ <b>Бюджет превышен!</b>"
        elif spent >= budget * 0.8:
            text += "\n\n⚠️ Использовано более 80% бюджета!"

    await callback.message.edit_text(text, parse_mode="HTML")
    await callback.message.answer("Главное меню:", reply_markup=main_menu_kb())
    await callback.answer()

# ========================= ADD GOAL FSM =========================

@dp.callback_query(F.data == "add_goal")
async def start_add_goal(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AddGoal.title)
    await callback.message.answer(
        "🎯 <b>Новая цель</b>\n\nВведите название цели:",
        parse_mode="HTML"
    )
    await callback.answer()

@dp.message(AddGoal.title)
async def goal_get_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await state.set_state(AddGoal.target)
    await message.answer(f"✅ Название: <b>{message.text}</b>\n\nВведите целевую сумму:", parse_mode="HTML")

@dp.message(AddGoal.target)
async def goal_get_target(message: Message, state: FSMContext):
    try:
        target = float(message.text.replace(",", ".").replace(" ", ""))
        if target <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму")
        return
    data = await state.get_data()
    title = data["title"]
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("INSERT INTO goals VALUES(NULL,?,?,?,0)", (message.from_user.id, title, target))
    conn.commit()
    conn.close()
    await state.clear()
    await message.answer(
        f"✅ Цель <b>{title}</b> на <b>{target:,.0f} ₽</b> создана!",
        parse_mode="HTML", reply_markup=main_menu_kb()
    )

# ========================= FUND GOAL FSM =========================

@dp.callback_query(F.data.startswith("fund_goal_"))
async def start_fund_goal(callback: CallbackQuery, state: FSMContext):
    goal_id = int(callback.data.split("_")[-1])
    await state.set_state(FundGoal.amount)
    await state.update_data(goal_id=goal_id)
    await callback.message.answer("💰 Введите сумму для пополнения цели:")
    await callback.answer()

@dp.message(FundGoal.amount)
async def fund_goal_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму")
        return
    data = await state.get_data()
    goal_id = data["goal_id"]
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE goals SET current=current+? WHERE id=? AND user_id=?",
                (amount, goal_id, message.from_user.id))
    cur.execute("SELECT title, target, current FROM goals WHERE id=?", (goal_id,))
    row = cur.fetchone()
    conn.commit()
    conn.close()
    await state.clear()
    if row:
        title, target, current = row
        bar = progress_bar(current, target)
        pct = round((current / target) * 100, 1) if target else 0
        extra = "\n\n🎉 <b>Цель достигнута! Поздравляем!</b>" if current >= target else ""
        await message.answer(
            f"✅ Цель <b>{title}</b> пополнена на <b>{amount:,.0f} ₽</b>!\n\n"
            f"{bar} {pct}%\nНакоплено: <b>{current:,.0f} ₽</b> из <b>{target:,.0f} ₽</b>{extra}",
            parse_mode="HTML", reply_markup=main_menu_kb()
        )
    else:
        await message.answer("❌ Цель не найдена.", reply_markup=main_menu_kb())

# ========================= 📊 CHARTS =========================

def make_pie_chart(uid) -> io.BytesIO | None:
    conn = get_conn()
    cur = conn.cursor()
    start = (datetime.now() - timedelta(days=30)).strftime("%Y-%m-%d")
    cur.execute("""
        SELECT category, SUM(amount) FROM expenses
        WHERE user_id=? AND date>=?
        GROUP BY category ORDER BY SUM(amount) DESC
    """, (uid, start))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        return None

    labels = [next((l for l, v in CATEGORIES if v == r[0]), r[0]) for r in rows]
    values = [r[1] for r in rows]

    colors = [
        "#FF6B6B", "#4ECDC4", "#45B7D1", "#96CEB4", "#FFEAA7",
        "#DDA0DD", "#98D8C8", "#F7DC6F", "#BB8FCE", "#85C1E9",
        "#F1948A", "#82E0AA", "#FAD7A0", "#AED6F1"
    ]

    fig, ax = plt.subplots(figsize=(8, 6), facecolor="#1a1a2e")
    ax.set_facecolor("#1a1a2e")

    wedges, texts, autotexts = ax.pie(
        values,
        labels=None,
        autopct=lambda p: f"{p:.1f}%" if p > 4 else "",
        colors=colors[:len(values)],
        startangle=140,
        wedgeprops=dict(linewidth=2, edgecolor="#1a1a2e"),
        pctdistance=0.82
    )

    for at in autotexts:
        at.set_fontsize(9)
        at.set_color("white")
        at.set_fontweight("bold")

    legend_labels = [f"{labels[i]}: {values[i]:,.0f} ₽" for i in range(len(labels))]
    patches = [mpatches.Patch(color=colors[i % len(colors)], label=legend_labels[i]) for i in range(len(labels))]
    ax.legend(handles=patches, loc="lower center", bbox_to_anchor=(0.5, -0.25),
              ncol=2, fontsize=8, framealpha=0.3, labelcolor="white",
              facecolor="#1a1a2e", edgecolor="gray")

    total = sum(values)
    ax.text(0, 0, f"{total:,.0f} ₽\nза 30 дней", ha="center", va="center",
            fontsize=11, color="white", fontweight="bold")

    ax.set_title("📊 Расходы по категориям", color="white", fontsize=14, fontweight="bold", pad=15)

    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", dpi=120, facecolor="#1a1a2e")
    plt.close()
    buf.seek(0)
    return buf


def make_bar_chart(uid) -> io.BytesIO | None:
    conn = get_conn()
    cur = conn.cursor()
    # Last 7 days bar chart
    days_data = []
    for i in range(6, -1, -1):
        d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
        cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=? AND date=?", (uid, d))
        val = cur.fetchone()[0] or 0
        days_data.append((d[5:], val))  # MM-DD format
    conn.close()

    if all(v == 0 for _, v in days_data):
        return None

    labels = [d for d, _ in days_data]
    values = [v for _, v in days_data]

    fig, ax = plt.subplots(figsize=(8, 5), facecolor="#1a1a2e")
    ax.set_facecolor("#16213e")

    bars = ax.bar(labels, values, color="#4ECDC4", edgecolor="#1a1a2e", linewidth=1.5, width=0.6)

    for bar, val in zip(bars, values):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(values) * 0.02,
                    f"{val:,.0f}", ha="center", va="bottom", color="white", fontsize=8, fontweight="bold")

    ax.set_title("📅 Расходы за последние 7 дней", color="white", fontsize=13, fontweight="bold")
    ax.set_xlabel("Дата", color="#aaa", fontsize=10)
    ax.set_ylabel("Сумма (₽)", color="#aaa", fontsize=10)
    ax.tick_params(colors="white")
    ax.spines[:].set_color("#333")
    ax.yaxis.grid(True, color="#333", linestyle="--", alpha=0.5)
    ax.set_axisbelow(True)

    buf = io.BytesIO()
    plt.savefig(buf, format="png", bbox_inches="tight", dpi=120, facecolor="#1a1a2e")
    plt.close()
    buf.seek(0)
    return buf

@dp.callback_query(F.data == "stats_chart")
async def stats_chart(callback: CallbackQuery):
    await callback.answer("Генерирую графики...")
    uid = callback.from_user.id

    pie = make_pie_chart(uid)
    bar = make_bar_chart(uid)

    if not pie and not bar:
        await callback.message.answer("📊 Недостаточно данных для построения графиков.")
        return

    if pie:
        await bot.send_photo(callback.from_user.id,
            BufferedInputFile(pie.read(), filename="chart_pie.png"),
            caption="🍕 Распределение расходов по категориям (30 дней)")
    if bar:
        await bot.send_photo(callback.from_user.id,
            BufferedInputFile(bar.read(), filename="chart_bar.png"),
            caption="📅 Расходы по дням (последние 7 дней)")

# ========================= 💱 CURRENCY CONVERTER =========================

@dp.message(F.text == "💱 Валюты")
async def menu_currency(message: Message, state: FSMContext):
    await state.set_state(CurrencyConvert.from_currency)
    await message.answer(
        "💱 <b>Конвертер валют</b>\n\nВыберите исходную валюту:",
        parse_mode="HTML",
        reply_markup=currencies_kb("from")
    )

@dp.callback_query(F.data.startswith("from_"), CurrencyConvert.from_currency)
async def currency_from(callback: CallbackQuery, state: FSMContext):
    from_cur = callback.data[5:]
    await state.update_data(from_currency=from_cur)
    await state.set_state(CurrencyConvert.amount)
    await callback.message.edit_text(
        f"✅ Из: <b>{from_cur}</b>\n\nВведите сумму:",
        parse_mode="HTML"
    )
    await callback.answer()

@dp.message(CurrencyConvert.amount)
async def currency_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму")
        return
    await state.update_data(amount=amount)
    await state.set_state(CurrencyConvert.to_currency)
    data = await state.get_data()
    await message.answer(
        f"✅ Сумма: <b>{amount:,.2f} {data['from_currency']}</b>\n\nВыберите валюту назначения:",
        parse_mode="HTML",
        reply_markup=currencies_kb("to")
    )

@dp.callback_query(F.data.startswith("to_"), CurrencyConvert.to_currency)
async def currency_to(callback: CallbackQuery, state: FSMContext):
    to_cur = callback.data[3:]
    data = await state.get_data()
    from_cur = data["from_currency"]
    amount = data["amount"]

    await state.clear()
    await callback.answer("Получаю курс...")

    try:
        async with aiohttp.ClientSession() as session:
            url = f"https://api.exchangerate-api.com/v4/latest/{from_cur}"
            async with session.get(url, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                if resp.status != 200:
                    raise Exception("API error")
                json_data = await resp.json()
                rate = json_data["rates"].get(to_cur)
                if not rate:
                    raise Exception("No rate")

        result = amount * rate
        await callback.message.edit_text(
            f"💱 <b>Результат конвертации</b>\n\n"
            f"<b>{amount:,.2f} {from_cur}</b>\n"
            f"= <b>{result:,.2f} {to_cur}</b>\n\n"
            f"Курс: 1 {from_cur} = {rate:.4f} {to_cur}\n"
            f"<i>Данные: exchangerate-api.com</i>",
            parse_mode="HTML"
        )
    except Exception:
        await callback.message.edit_text(
            "❌ Не удалось получить курс валют.\nПроверьте интернет-соединение."
        )

    await callback.message.answer("Главное меню:", reply_markup=main_menu_kb())

# ========================= 📤 EXPORT EXCEL =========================

@dp.message(F.text == "📤 Экспорт Excel")
async def export_excel(message: Message):
    uid = message.from_user.id
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("SELECT amount, source, date FROM incomes WHERE user_id=? ORDER BY date DESC", (uid,))
    incomes = cur.fetchall()

    cur.execute("SELECT amount, category, date FROM expenses WHERE user_id=? ORDER BY date DESC", (uid,))
    expenses = cur.fetchall()

    cur.execute("SELECT title, target, current FROM goals WHERE user_id=?", (uid,))
    goals = cur.fetchall()

    conn.close()

    if not incomes and not expenses:
        await message.answer("❌ Нет данных для экспорта.")
        return

    wb = openpyxl.Workbook()

    # === Styles ===
    header_font = Font(bold=True, color="FFFFFF", size=12)
    income_fill = PatternFill("solid", fgColor="27AE60")
    expense_fill = PatternFill("solid", fgColor="E74C3C")
    goal_fill = PatternFill("solid", fgColor="2980B9")
    summary_fill = PatternFill("solid", fgColor="8E44AD")
    center = Alignment(horizontal="center", vertical="center")

    def style_header(ws, row, col, text, fill):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = center

    # === Лист 1: Доходы ===
    ws1 = wb.active
    ws1.title = "💰 Доходы"
    headers = ["#", "Сумма (₽)", "Источник", "Дата"]
    for i, h in enumerate(headers, 1):
        style_header(ws1, 1, i, h, income_fill)
        ws1.column_dimensions[chr(64 + i)].width = [5, 15, 25, 15][i - 1]

    total_income = 0
    for idx, (amount, source, date) in enumerate(incomes, 2):
        ws1.cell(row=idx, column=1, value=idx - 1)
        ws1.cell(row=idx, column=2, value=amount)
        ws1.cell(row=idx, column=3, value=source or "—")
        ws1.cell(row=idx, column=4, value=date)
        total_income += amount

    last = len(incomes) + 2
    ws1.cell(row=last, column=2, value=f"Итого: {total_income:,.0f} ₽").font = Font(bold=True)

    # === Лист 2: Расходы ===
    ws2 = wb.create_sheet("💸 Расходы")
    headers2 = ["#", "Сумма (₽)", "Категория", "Дата"]
    for i, h in enumerate(headers2, 1):
        style_header(ws2, 1, i, h, expense_fill)
        ws2.column_dimensions[chr(64 + i)].width = [5, 15, 20, 15][i - 1]

    total_expense = 0
    for idx, (amount, category, date) in enumerate(expenses, 2):
        cat_label = next((l for l, v in CATEGORIES if v == category), category)
        ws2.cell(row=idx, column=1, value=idx - 1)
        ws2.cell(row=idx, column=2, value=amount)
        ws2.cell(row=idx, column=3, value=cat_label)
        ws2.cell(row=idx, column=4, value=date)
        total_expense += amount

    last2 = len(expenses) + 2
    ws2.cell(row=last2, column=2, value=f"Итого: {total_expense:,.0f} ₽").font = Font(bold=True)

    # === Лист 3: Цели ===
    ws3 = wb.create_sheet("🎯 Цели")
    headers3 = ["Название", "Цель (₽)", "Накоплено (₽)", "Прогресс (%)"]
    for i, h in enumerate(headers3, 1):
        style_header(ws3, 1, i, h, goal_fill)
        ws3.column_dimensions[chr(64 + i)].width = [25, 15, 18, 15][i - 1]

    for idx, (title, target, current) in enumerate(goals, 2):
        pct = round((current / target) * 100, 1) if target else 0
        ws3.cell(row=idx, column=1, value=title)
        ws3.cell(row=idx, column=2, value=target)
        ws3.cell(row=idx, column=3, value=current)
        ws3.cell(row=idx, column=4, value=pct)

    # === Лист 4: Сводка ===
    ws4 = wb.create_sheet("📊 Сводка")
    style_header(ws4, 1, 1, "Показатель", summary_fill)
    style_header(ws4, 1, 2, "Значение", summary_fill)
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 20

    balance = total_income - total_expense
    summary_data = [
        ("Всего доходов", f"{total_income:,.0f} ₽"),
        ("Всего расходов", f"{total_expense:,.0f} ₽"),
        ("Баланс", f"{balance:,.0f} ₽"),
        ("Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M")),
    ]
    for i, (k, v) in enumerate(summary_data, 2):
        ws4.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws4.cell(row=i, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"finance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=filename),
        caption=(
            f"📤 <b>Экспорт финансов</b>\n\n"
            f"💰 Доходов: {len(incomes)} записей\n"
            f"💸 Расходов: {len(expenses)} записей\n"
            f"🎯 Целей: {len(goals)}\n"
            f"💼 Баланс: <b>{balance:,.0f} ₽</b>"
        ),
        parse_mode="HTML"
    )

# ========================= 💰 BUDGETS =========================

@dp.callback_query(F.data == "budgets_view")
async def budgets_view(callback: CallbackQuery):
    uid = callback.from_user.id
    month = current_month()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT category, amount FROM budgets WHERE user_id=? AND month=?", (uid, month))
    budgets = {row[0]: row[1] for row in cur.fetchall()}
    conn.close()

    if not budgets:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Установить бюджет", callback_data="set_budget")]
        ])
        await callback.message.answer("💰 Бюджеты на этот месяц не установлены.", reply_markup=kb)
        await callback.answer()
        return

    conn = get_conn()
    cur = conn.cursor()
    text = f"💰 <b>Бюджеты на {month}:</b>\n\n"
    for cat, budget in budgets.items():
        cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=? AND category=? AND date LIKE ?",
                    (uid, cat, f"{month}%"))
        spent = cur.fetchone()[0] or 0
        pct = round((spent / budget) * 100)
        bar = progress_bar(spent, budget)
        cat_label = next((l for l, v in CATEGORIES if v == cat), cat)
        status = "⚠️" if spent >= budget else ("🟡" if spent >= budget * 0.8 else "🟢")
        text += f"{status} <b>{cat_label}</b>\n{bar} {pct}%\n{spent:,.0f} / {budget:,.0f} ₽\n\n"
    conn.close()

    kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить/изменить бюджет", callback_data="set_budget")]
    ])
    await callback.message.answer(text, parse_mode="HTML", reply_markup=kb)
    await callback.answer()

@dp.callback_query(F.data == "set_budget")
async def start_set_budget(callback: CallbackQuery, state: FSMContext):
    await state.set_state(SetBudget.category)
    await callback.message.answer(
        "💰 <b>Установка бюджета</b>\n\nВыберите категорию:",
        parse_mode="HTML",
        reply_markup=budget_categories_kb()
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("budget_cat_"), SetBudget.category)
async def budget_get_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data[11:]
    await state.update_data(category=category)
    await state.set_state(SetBudget.amount)
    cat_label = next((l for l, v in CATEGORIES if v == category), category)
    await callback.message.edit_text(
        f"✅ Категория: <b>{cat_label}</b>\n\nВведите лимит бюджета на месяц (₽):",
        parse_mode="HTML"
    )
    await callback.answer()

@dp.message(SetBudget.amount)
async def budget_get_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму")
        return
    data = await state.get_data()
    category = data["category"]
    uid = message.from_user.id
    month = current_month()

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO budgets(user_id, category, amount, month)
        VALUES(?,?,?,?)
        ON CONFLICT(user_id, category, month) DO UPDATE SET amount=excluded.amount
    """, (uid, category, amount, month))
    conn.commit()
    conn.close()

    await state.clear()
    cat_label = next((l for l, v in CATEGORIES if v == category), category)
    await message.answer(
        f"✅ Бюджет для <b>{cat_label}</b> установлен: <b>{amount:,.0f} ₽/мес</b>",
        parse_mode="HTML",
        reply_markup=main_menu_kb()
    )

# ========================= 🔔 DAILY REMINDERS =========================

@dp.callback_query(F.data == "toggle_reminder")
async def toggle_reminder(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT user_id FROM reminder_users WHERE user_id=?", (uid,))
    exists = cur.fetchone()

    if exists:
        cur.execute("DELETE FROM reminder_users WHERE user_id=?", (uid,))
        text = "🔕 Ежедневные напоминания <b>отключены</b>."
        reminders_on = False
    else:
        cur.execute("INSERT INTO reminder_users VALUES(?)", (uid,))
        text = f"🔔 Ежедневные напоминания <b>включены</b>!\nКаждый день в <b>{REMINDER_HOUR:02d}:{REMINDER_MINUTE:02d}</b> я напомню записать расходы."
        reminders_on = True

    conn.commit()
    conn.close()

    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=management_kb(reminders_on))
    await callback.answer()

async def daily_reminder_task():
    """Background task: sends reminder every day at REMINDER_HOUR:REMINDER_MINUTE"""
    while True:
        now = datetime.now()
        target = now.replace(hour=REMINDER_HOUR, minute=REMINDER_MINUTE, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        wait_seconds = (target - now).total_seconds()
        await asyncio.sleep(wait_seconds)

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT user_id FROM reminder_users")
        users = [row[0] for row in cur.fetchall()]
        conn.close()

        today = today_date()
        for uid in users:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=? AND date=?", (uid, today))
            spent_today = cur.fetchone()[0] or 0
            conn.close()

            if spent_today > 0:
                msg = (
                    f"🔔 <b>Ежедневный отчёт</b>\n\n"
                    f"Сегодня ты потратил: <b>{spent_today:,.0f} ₽</b>\n"
                    f"Не забудь записать все расходы!"
                )
            else:
                msg = (
                    "🔔 <b>Напоминание</b>\n\n"
                    "Ты ещё не записал расходы сегодня.\n"
                    "Нажми <b>➖ Расход</b> чтобы добавить!"
                )

            try:
                await bot.send_message(uid, msg, parse_mode="HTML")
            except Exception:
                pass  # User blocked bot etc.

# ========================= STATISTICS CALLBACKS =========================

@dp.callback_query(F.data.startswith("stats_"))
async def stats_callback(callback: CallbackQuery):
    uid = callback.from_user.id
    period = callback.data[6:]

    if period == "chart":
        return  # handled separately above

    if period == "today":
        today = today_date()
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT SUM(amount) FROM incomes WHERE user_id=? AND date=?", (uid, today))
        income = cur.fetchone()[0] or 0
        cur.execute("SELECT SUM(amount) FROM expenses WHERE user_id=? AND date=?", (uid, today))
        expense = cur.fetchone()[0] or 0
        conn.close()
        title = "📅 Сегодня"
    elif period == "week":
        income, expense = get_period_stats(uid, 7)
        title = "📆 За 7 дней"
    elif period == "month":
        income, expense = get_period_stats(uid, 30)
        title = "🗓 За 30 дней"
    elif period == "all":
        income, expense = get_period_stats(uid)
        title = "📈 За всё время"
    elif period == "top":
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT category, SUM(amount) as total
            FROM expenses WHERE user_id=?
            GROUP BY category ORDER BY total DESC LIMIT 5
        """, (uid,))
        rows = cur.fetchall()
        conn.close()
        if not rows:
            await callback.answer("Нет данных о расходах", show_alert=True)
            return
        text = "🏆 <b>Топ категорий расходов:</b>\n\n"
        medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
        max_amount = rows[0][1]
        for i, (cat, total) in enumerate(rows):
            bar = progress_bar(total, max_amount, 8)
            cat_label = next((l for l, v in CATEGORIES if v == cat), cat)
            text += f"{medals[i]} {cat_label}\n{bar} <b>{total:,.0f} ₽</b>\n\n"
        await callback.message.edit_text(text, parse_mode="HTML", reply_markup=stats_kb())
        await callback.answer()
        return
    elif period == "avg":
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT AVG(amount) FROM expenses WHERE user_id=?", (uid,))
        avg = cur.fetchone()[0] or 0
        cur.execute("SELECT COUNT(*) FROM expenses WHERE user_id=?", (uid,))
        count = cur.fetchone()[0] or 0
        conn.close()
        await callback.message.edit_text(
            f"📉 <b>Средний расход</b>\n\nСумма: <b>{avg:,.0f} ₽</b>\nВсего операций: <b>{count}</b>",
            parse_mode="HTML", reply_markup=stats_kb()
        )
        await callback.answer()
        return
    else:
        return

    balance = income - expense
    balance_emoji = "🟢" if balance >= 0 else "🔴"
    await callback.message.edit_text(
        f"📊 <b>{title}</b>\n\n"
        f"💰 Доходы: <b>{income:,.0f} ₽</b>\n"
        f"💸 Расходы: <b>{expense:,.0f} ₽</b>\n"
        f"{'─' * 20}\n"
        f"{balance_emoji} Баланс: <b>{balance:,.0f} ₽</b>",
        parse_mode="HTML", reply_markup=stats_kb()
    )
    await callback.answer()

# ========================= HISTORY =========================

@dp.callback_query(F.data == "hist_incomes")
async def hist_incomes(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, amount, source, date FROM incomes WHERE user_id=? ORDER BY id DESC LIMIT 10", (uid,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await callback.answer("Нет доходов", show_alert=True)
        return
    text = "💰 <b>Последние доходы:</b>\n\n"
    for id_, amount, source, date in rows:
        src = f" ({source})" if source else ""
        text += f"[#{id_}] <b>{amount:,.0f} ₽</b>{src} — {date}\n"
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=history_kb())
    await callback.answer()

@dp.callback_query(F.data == "hist_expenses")
async def hist_expenses(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, category, amount, date FROM expenses WHERE user_id=? ORDER BY id DESC LIMIT 10", (uid,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await callback.answer("Нет расходов", show_alert=True)
        return
    text = "💸 <b>Последние расходы:</b>\n\n"
    for id_, category, amount, date in rows:
        cat_label = next((l for l, v in CATEGORIES if v == category), category)
        text += f"[#{id_}] {cat_label} — <b>{amount:,.0f} ₽</b> ({date})\n"
    await callback.message.edit_text(text, parse_mode="HTML", reply_markup=history_kb())
    await callback.answer()

# ========================= DELETE FLOWS =========================

@dp.callback_query(F.data == "del_income_list")
async def del_income_list(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, amount, date FROM incomes WHERE user_id=? ORDER BY id DESC LIMIT 10", (uid,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await callback.answer("Нет доходов для удаления", show_alert=True)
        return
    buttons = [[InlineKeyboardButton(
        text=f"❌ #{id_} — {amount:,.0f} ₽ ({date})",
        callback_data=f"confirm_del_income_{id_}"
    )] for id_, amount, date in rows]
    await callback.message.edit_text("Выберите доход для удаления:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@dp.callback_query(F.data.startswith("confirm_del_income_"))
async def confirm_del_income(callback: CallbackQuery):
    income_id = int(callback.data.split("_")[-1])
    await callback.message.edit_text(f"Удалить доход #{income_id}?",
        reply_markup=confirm_kb(f"do_del_income_{income_id}"))
    await callback.answer()

@dp.callback_query(F.data.startswith("do_del_income_"))
async def do_del_income(callback: CallbackQuery):
    income_id = int(callback.data.split("_")[-1])
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM incomes WHERE id=? AND user_id=?", (income_id, callback.from_user.id))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    await callback.message.edit_text("✅ Доход удалён." if deleted else "❌ Запись не найдена.")
    await callback.answer()

@dp.callback_query(F.data == "del_expense_list")
async def del_expense_list(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, category, amount, date FROM expenses WHERE user_id=? ORDER BY id DESC LIMIT 10", (uid,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await callback.answer("Нет расходов для удаления", show_alert=True)
        return
    buttons = [[InlineKeyboardButton(
        text=f"❌ #{id_} {next((l for l, v in CATEGORIES if v == cat), cat)} — {amount:,.0f} ₽ ({date})",
        callback_data=f"confirm_del_expense_{id_}"
    )] for id_, cat, amount, date in rows]
    await callback.message.edit_text("Выберите расход для удаления:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@dp.callback_query(F.data.startswith("confirm_del_expense_"))
async def confirm_del_expense(callback: CallbackQuery):
    expense_id = int(callback.data.split("_")[-1])
    await callback.message.edit_text(f"Удалить расход #{expense_id}?",
        reply_markup=confirm_kb(f"do_del_expense_{expense_id}"))
    await callback.answer()

@dp.callback_query(F.data.startswith("do_del_expense_"))
async def do_del_expense(callback: CallbackQuery):
    expense_id = int(callback.data.split("_")[-1])
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM expenses WHERE id=? AND user_id=?", (expense_id, callback.from_user.id))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    await callback.message.edit_text("✅ Расход удалён." if deleted else "❌ Запись не найдена.")
    await callback.answer()

@dp.callback_query(F.data == "del_goal_list")
async def del_goal_list(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, title, current, target FROM goals WHERE user_id=?", (uid,))
    rows = cur.fetchall()
    conn.close()
    if not rows:
        await callback.answer("Нет целей для удаления", show_alert=True)
        return
    buttons = [[InlineKeyboardButton(
        text=f"❌ {title} ({current:,.0f}/{target:,.0f} ₽)",
        callback_data=f"confirm_del_goal_{id_}"
    )] for id_, title, current, target in rows]
    await callback.message.edit_text("Выберите цель для удаления:", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))
    await callback.answer()

@dp.callback_query(F.data.startswith("confirm_del_goal_"))
async def confirm_del_goal(callback: CallbackQuery):
    goal_id = int(callback.data.split("_")[-1])
    await callback.message.edit_text(f"Удалить цель #{goal_id}?",
        reply_markup=confirm_kb(f"do_del_goal_{goal_id}"))
    await callback.answer()

@dp.callback_query(F.data.startswith("do_del_goal_"))
async def do_del_goal(callback: CallbackQuery):
    goal_id = int(callback.data.split("_")[-1])
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM goals WHERE id=? AND user_id=?", (goal_id, callback.from_user.id))
    deleted = cur.rowcount
    conn.commit()
    conn.close()
    await callback.message.edit_text("✅ Цель удалена." if deleted else "❌ Цель не найдена.")
    await callback.answer()

# ========================= CLEAR ALL =========================

@dp.callback_query(F.data == "clear_all")
async def clear_all_confirm(callback: CallbackQuery):
    await callback.message.edit_text(
        "⚠️ <b>Внимание!</b>\n\nВы собираетесь удалить <b>ВСЕ</b> ваши данные.\nЭто действие необратимо!",
        parse_mode="HTML", reply_markup=confirm_kb("do_clear_all")
    )
    await callback.answer()

@dp.callback_query(F.data == "do_clear_all")
async def do_clear_all(callback: CallbackQuery):
    uid = callback.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM expenses WHERE user_id=?", (uid,))
    cur.execute("DELETE FROM incomes WHERE user_id=?", (uid,))
    cur.execute("DELETE FROM goals WHERE user_id=?", (uid,))
    cur.execute("DELETE FROM budgets WHERE user_id=?", (uid,))
    conn.commit()
    conn.close()
    await callback.message.edit_text("✅ Все данные удалены.")
    await callback.answer()

@dp.callback_query(F.data == "cancel_del")
async def cancel_del(callback: CallbackQuery):
    await callback.message.edit_text("❌ Отменено.")
    await callback.answer()

@dp.message(Command("cancel"))
async def cmd_cancel(message: Message, state: FSMContext):
    await state.clear()
    await message.answer("❌ Действие отменено.", reply_markup=main_menu_kb())

# ========================= 🔄 RECURRING EXPENSES =========================

@dp.message(F.text == "🔄 Регулярные")
async def menu_recurring(message: Message):
    uid = message.from_user.id
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, title, amount, category, day_of_month FROM recurring WHERE user_id=?", (uid,))
    rows = cur.fetchall()
    conn.close()

    if not rows:
        kb = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Добавить регулярный расход", callback_data="add_recurring")]
        ])
        await message.answer(
            "🔄 <b>Регулярные расходы</b>\n\n"
            "Здесь можно настроить расходы, которые повторяются каждый месяц.\n"
            "Например: аренда, Netflix, абонемент в зал.\n\n"
            "Бот будет автоматически добавлять их в указанный день месяца.",
            parse_mode="HTML",
            reply_markup=kb
        )
        return

    text = "🔄 <b>Регулярные расходы:</b>\n\n"
    buttons = []
    for id_, title, amount, category, day in rows:
        cat_label = next((l for l, v in CATEGORIES if v == category), category)
        text += f"• <b>{title}</b> — {amount:,.0f} ₽ ({cat_label}), каждое <b>{day}-е</b> число\n"
        buttons.append([InlineKeyboardButton(
            text=f"🗑 Удалить: {title}",
            callback_data=f"del_recurring_{id_}"
        )])

    buttons.append([InlineKeyboardButton(text="➕ Добавить ещё", callback_data="add_recurring")])
    await message.answer(text, parse_mode="HTML", reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons))

@dp.callback_query(F.data == "add_recurring")
async def start_add_recurring(callback: CallbackQuery, state: FSMContext):
    await state.set_state(AddRecurring.title)
    await callback.message.answer(
        "🔄 <b>Новый регулярный расход</b>\n\nВведите название (например: Netflix, Аренда, Зал):",
        parse_mode="HTML"
    )
    await callback.answer()

@dp.message(AddRecurring.title)
async def recurring_title(message: Message, state: FSMContext):
    await state.update_data(title=message.text)
    await state.set_state(AddRecurring.amount)
    await message.answer(f"✅ Название: <b>{message.text}</b>\n\nВведите сумму:", parse_mode="HTML")

@dp.message(AddRecurring.amount)
async def recurring_amount(message: Message, state: FSMContext):
    try:
        amount = float(message.text.replace(",", ".").replace(" ", ""))
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите корректную сумму")
        return
    await state.update_data(amount=amount)
    await state.set_state(AddRecurring.category)
    await message.answer(
        f"✅ Сумма: <b>{amount:,.0f} ₽</b>\n\nВыберите категорию:",
        parse_mode="HTML",
        reply_markup=categories_kb()
    )

@dp.callback_query(F.data.startswith("cat_"), AddRecurring.category)
async def recurring_category(callback: CallbackQuery, state: FSMContext):
    category = callback.data[4:]
    await state.update_data(category=category)
    await state.set_state(AddRecurring.day)

    days_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=str(d), callback_data=f"rec_day_{d}") for d in range(1, 6)],
        [InlineKeyboardButton(text=str(d), callback_data=f"rec_day_{d}") for d in range(6, 11)],
        [InlineKeyboardButton(text=str(d), callback_data=f"rec_day_{d}") for d in range(11, 16)],
        [InlineKeyboardButton(text=str(d), callback_data=f"rec_day_{d}") for d in range(16, 21)],
        [InlineKeyboardButton(text=str(d), callback_data=f"rec_day_{d}") for d in range(21, 26)],
        [InlineKeyboardButton(text=str(d), callback_data=f"rec_day_{d}") for d in range(26, 29)],
    ])
    cat_label = next((l for l, v in CATEGORIES if v == category), category)
    await callback.message.edit_text(
        f"✅ Категория: <b>{cat_label}</b>\n\nКакого числа каждого месяца списывать?",
        parse_mode="HTML",
        reply_markup=days_kb
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("rec_day_"), AddRecurring.day)
async def recurring_day(callback: CallbackQuery, state: FSMContext):
    day = int(callback.data.split("_")[-1])
    data = await state.get_data()
    uid = callback.from_user.id

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO recurring(user_id, title, amount, category, day_of_month, last_applied) VALUES(?,?,?,?,?,?)",
        (uid, data["title"], data["amount"], data["category"], day, "")
    )
    conn.commit()
    conn.close()

    await state.clear()
    cat_label = next((l for l, v in CATEGORIES if v == data["category"]), data["category"])
    await callback.message.edit_text(
        f"✅ Регулярный расход добавлен!\n\n"
        f"📌 <b>{data['title']}</b>\n"
        f"💰 {data['amount']:,.0f} ₽ — {cat_label}\n"
        f"📅 Каждое <b>{day}-е</b> число месяца\n\n"
        f"Бот будет автоматически добавлять этот расход в нужный день.",
        parse_mode="HTML"
    )
    await callback.message.answer("Главное меню:", reply_markup=main_menu_kb())
    await callback.answer()

@dp.callback_query(F.data.startswith("del_recurring_"))
async def del_recurring(callback: CallbackQuery):
    rec_id = int(callback.data.split("_")[-1])
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM recurring WHERE id=? AND user_id=?", (rec_id, callback.from_user.id))
    conn.commit()
    conn.close()
    await callback.answer("✅ Удалено", show_alert=True)
    await callback.message.delete()

async def apply_recurring_task():
    """Background task: applies recurring expenses on their day of month"""
    while True:
        now = datetime.now()
        # Run at 08:00 every day
        target = now.replace(hour=8, minute=0, second=0, microsecond=0)
        if now >= target:
            target += timedelta(days=1)
        await asyncio.sleep((target - now).total_seconds())

        today = today_date()
        day = datetime.now().day

        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, user_id, title, amount, category FROM recurring WHERE day_of_month=?", (day,))
        rows = cur.fetchall()

        applied = []
        for id_, uid, title, amount, category in rows:
            # Check if already applied this month
            month_prefix = datetime.now().strftime("%Y-%m")
            cur.execute("SELECT last_applied FROM recurring WHERE id=?", (id_,))
            last = cur.fetchone()[0] or ""
            if last.startswith(month_prefix):
                continue

            cur.execute("INSERT INTO expenses VALUES(NULL,?,?,?,?)", (uid, amount, category, today))
            cur.execute("UPDATE recurring SET last_applied=? WHERE id=?", (today, id_))
            applied.append((uid, title, amount))

        conn.commit()
        conn.close()

        # Notify users
        for uid, title, amount in applied:
            try:
                await bot.send_message(
                    uid,
                    f"🔄 <b>Регулярный расход списан</b>\n\n"
                    f"📌 {title} — <b>{amount:,.0f} ₽</b>\n"
                    f"Добавлен автоматически.",
                    parse_mode="HTML"
                )
            except Exception:
                pass

# ========================= MAIN =========================

async def main():
    init_db()
    print("🤖 Финансовый бот v4 запущен!")
    asyncio.create_task(daily_reminder_task())
    asyncio.create_task(apply_recurring_task())
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())